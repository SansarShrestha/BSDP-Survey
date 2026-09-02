import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
import json
import os

# Page config
st.set_page_config(
    page_title="Tool 1 — Stage 1: Rapid Innovation Screening",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        font-size: 2rem;
        font-weight: bold;
        color: #1f4e79;
        margin-bottom: 0.5rem;
    }
    .sub-header {
        font-size: 1.1rem;
        color: #555;
        margin-bottom: 2rem;
    }
    .stButton>button {
        background-color: #1f4e79;
        color: white;
        border-radius: 5px;
        padding: 0.5rem 2rem;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 5px;
        padding: 1rem;
        color: #155724;
    }
    .warning-box {
        background-color: #fff3cd;
        border: 1px solid #ffeaa7;
        border-radius: 5px;
        padding: 1rem;
        color: #856404;
    }
    .metric-card {
        background-color: #f8f9fa;
        border-radius: 10px;
        padding: 1.5rem;
        text-align: center;
        border: 1px solid #dee2e6;
    }
    .metric-value {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f4e79;
    }
    .metric-label {
        font-size: 0.9rem;
        color: #666;
        margin-top: 0.5rem;
    }
    .eligibility-yes {
        color: #28a745;
        font-weight: bold;
    }
    .eligibility-no {
        color: #dc3545;
        font-weight: bold;
    }
    .result-satisfactory {
        color: #28a745;
        font-weight: bold;
        font-size: 1.1rem;
    }
    .result-nonsatisfactory {
        color: #dc3545;
        font-weight: bold;
        font-size: 1.1rem;
    }
</style>
""", unsafe_allow_html=True)

# Initialize session state
if 'entries' not in st.session_state:
    st.session_state.entries = []
if 'threshold' not in st.session_state:
    st.session_state.threshold = 0
if 'admin_logged_in' not in st.session_state:
    st.session_state.admin_logged_in = False

# Constants
ELIGIBILITY_QUESTIONS = {
    "legal_compliance": "1. Legal compliance",
    "sector_eligibility": "2. Sector eligibility",
    "innovation_relevance": "3. Innovation relevance",
    "commitment_sme": "4. Commitment of SME",
    "environmental_compliance": "5. Environmental compliance",
    "hr_capacity": "6. HR Capacity (Technical/Business)",
    "benefit_women": "7. Benefit to women/marginalized groups"
}

INNOVATION_TYPES = ["Product Innovation", "Process Innovation", "Marketing Innovation", 
                    "Business Model Innovation", "Not clear yet", "None"]

SECTORS = ["Agro-processing", "Manufacturing", "Retail", "Services", "IT/Tech", 
           "Herbs/Production", "Automobile", "Tourism", "Construction", "Other"]

BANKS = ["Global IME Bank", "Rastriya Banijya Bank", "Nabil Bank", "NIC Asia Bank", 
         "Planeteer Innovation BDSP", "Other"]

# Helper functions
def calculate_result(answers, threshold):
    no_count = sum(1 for v in answers.values() if v == "No")
    if no_count <= threshold:
        return "Satisfactory", "Recommended for Stage 2 Assessment", no_count
    else:
        return "Non-Satisfactory", "Not Recommended", no_count

def create_excel_with_formulas(entries, threshold):
    wb = Workbook()

    # Styles
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    header_font = Font(bold=True, size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Instructions sheet
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    ws_instr['A1'] = "TOOL 1 — STAGE 1: RAPID INNOVATION ELIGIBILITY SCREENING TOOL"
    ws_instr['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws_instr['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_instr['A3'] = "Maximum number of 'No' answers still considered Satisfactory:"
    ws_instr['B3'] = threshold
    ws_instr['B3'].fill = yellow_fill
    ws_instr.column_dimensions['A'].width = 55
    ws_instr.column_dimensions['B'].width = 15

    # Screening Register sheet
    ws_reg = wb.create_sheet("Screening Register")
    headers = ["SN", "Screening Date", "SME Name", "District / Location", "Sector",
               "Loan Amount Applied (NPR)", "BDSP / Partner Bank Name", "Screener Name",
               "1. Legal compliance", "2. Sector eligibility", "3. Innovation relevance",
               "4. Commitment of SME", "5. Environmental compliance", "6. HR Capacity",
               "7. Benefit to women/marginalized", "Likely Innovation Type",
               "No. of 'No'", "Result", "Action", "Remarks"]

    for c_idx, header in enumerate(headers, 1):
        cell = ws_reg.cell(row=1, column=c_idx, value=header)
        cell.font = header_font
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    col_widths = [5, 12, 22, 15, 15, 18, 22, 15, 12, 12, 12, 12, 12, 12, 12, 18, 10, 15, 28, 35]
    for i, width in enumerate(col_widths, 1):
        ws_reg.column_dimensions[ws_reg.cell(row=1, column=i).column_letter].width = width

    # Data validation
    yes_no_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws_reg.add_data_validation(yes_no_validation)
    yes_no_validation.add('I2:O1000')

    innovation_validation = DataValidation(type="list", 
        formula1='"Product Innovation,Process Innovation,Marketing Innovation,Business Model Innovation,Not clear yet,None"', 
        allow_blank=True)
    ws_reg.add_data_validation(innovation_validation)
    innovation_validation.add('P2:P1000')

    # Add entries
    for idx, entry in enumerate(entries, 1):
        row = idx + 1
        ws_reg.cell(row=row, column=1, value=idx)
        ws_reg.cell(row=row, column=2, value=entry.get('screening_date', ''))
        ws_reg.cell(row=row, column=3, value=entry.get('sme_name', ''))
        ws_reg.cell(row=row, column=4, value=entry.get('district', ''))
        ws_reg.cell(row=row, column=5, value=entry.get('sector', ''))
        ws_reg.cell(row=row, column=6, value=entry.get('loan_amount', ''))
        ws_reg.cell(row=row, column=7, value=entry.get('bank_name', ''))
        ws_reg.cell(row=row, column=8, value=entry.get('screener_name', ''))
        ws_reg.cell(row=row, column=9, value=entry.get('legal_compliance', ''))
        ws_reg.cell(row=row, column=10, value=entry.get('sector_eligibility', ''))
        ws_reg.cell(row=row, column=11, value=entry.get('innovation_relevance', ''))
        ws_reg.cell(row=row, column=12, value=entry.get('commitment_sme', ''))
        ws_reg.cell(row=row, column=13, value=entry.get('environmental_compliance', ''))
        ws_reg.cell(row=row, column=14, value=entry.get('hr_capacity', ''))
        ws_reg.cell(row=row, column=15, value=entry.get('benefit_women', ''))
        ws_reg.cell(row=row, column=16, value=entry.get('innovation_type', ''))
        ws_reg.cell(row=row, column=20, value=entry.get('remarks', ''))

        # Yellow fill for input cells
        for col in range(1, 21):
            if col not in [17, 18, 19]:
                ws_reg.cell(row=row, column=col).fill = yellow_fill
            ws_reg.cell(row=row, column=col).border = border
            ws_reg.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='center')

        # Formula cells
        ws_reg.cell(row=row, column=17, value=f'=COUNTIF(I{row}:O{row},"No")')
        ws_reg.cell(row=row, column=17).fill = grey_fill if row % 2 == 0 else white_fill
        ws_reg.cell(row=row, column=17).border = border

        ws_reg.cell(row=row, column=18, value=f'=IF(Q{row}<=$Instructions.$B$3,"Satisfactory","Non-Satisfactory")')
        ws_reg.cell(row=row, column=18).fill = grey_fill if row % 2 == 0 else white_fill
        ws_reg.cell(row=row, column=18).border = border

        ws_reg.cell(row=row, column=19, value=f'=IF(R{row}="Satisfactory","Recommended for Stage 2 Assessment","Not Recommended")')
        ws_reg.cell(row=row, column=19).fill = grey_fill if row % 2 == 0 else white_fill
        ws_reg.cell(row=row, column=19).border = border

    # Dashboard sheet
    ws_dash = wb.create_sheet("Dashboard")
    ws_dash['A1'] = "STAGE 1 SCREENING — PIPELINE DASHBOARD"
    ws_dash['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws_dash['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    ws_dash['A3'] = "Summary"
    ws_dash['B3'] = "Count"
    ws_dash['A3'].font = header_font
    ws_dash['B3'].font = header_font

    ws_dash['A4'] = "Total applications screened"
    ws_dash['B4'] = f'=COUNTA('Screening Register'!A2:A{len(entries)+1})'

    ws_dash['A5'] = "Satisfactory — Recommended for Stage 2"
    ws_dash['B5'] = '=COUNTIF('Screening Register'!R:R,"Satisfactory")'

    ws_dash['A6'] = "Non-Satisfactory — Not Recommended"
    ws_dash['B6'] = '=COUNTIF('Screening Register'!R:R,"Non-Satisfactory")'

    ws_dash['A7'] = "% Recommended for Stage 2"
    ws_dash['B7'] = '=IF(B4>0,B5/B4,0)'
    ws_dash['B7'].number_format = '0.0%'

    ws_dash['A9'] = "By Likely Innovation Type"
    ws_dash['A9'].font = header_font

    for i, itype in enumerate(INNOVATION_TYPES, 10):
        ws_dash[f'A{i}'] = itype
        ws_dash[f'B{i}'] = f'=COUNTIF('Screening Register'!P:P,"{itype}")'

    for row in range(3, 16):
        for col in ['A', 'B']:
            ws_dash[f'{col}{row}'].border = border

    ws_dash.column_dimensions['A'].width = 40
    ws_dash.column_dimensions['B'].width = 15

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def get_dashboard_stats(entries, threshold):
    total = len(entries)
    satisfactory = 0
    innovation_counts = {t: 0 for t in INNOVATION_TYPES}

    for entry in entries:
        answers = {k: entry[k] for k in ELIGIBILITY_QUESTIONS.keys()}
        result, _, _ = calculate_result(answers, threshold)
        if result == "Satisfactory":
            satisfactory += 1
        itype = entry.get('innovation_type', 'None')
        if itype in innovation_counts:
            innovation_counts[itype] += 1

    non_satisfactory = total - satisfactory
    pct = (satisfactory / total * 100) if total > 0 else 0

    return total, satisfactory, non_satisfactory, pct, innovation_counts

# Sidebar navigation
st.sidebar.title("📋 Navigation")
page = st.sidebar.radio("Go to", ["🏠 Home", "📝 New Screening", "📊 Admin Dashboard", "📥 Download Data"])

# Home Page
if page == "🏠 Home":
    st.markdown('<div class="main-header">TOOL 1 — STAGE 1: Rapid Innovation Eligibility Screening</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-header">A quick filter to decide whether an SME loan application is worth carrying forward for full innovation assessment (Tool 2).</div>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown('<div class="metric-card"><div class="metric-value">7</div><div class="metric-label">Eligibility Parameters</div></div>', unsafe_allow_html=True)
    with col2:
        st.markdown('<div class="metric-card"><div class="metric-value">5</div><div class="metric-label">Innovation Types</div></div>', unsafe_allow_html=True)
    with col3:
        st.markdown('<div class="metric-card"><div class="metric-value">3</div><div class="metric-label">Auto-Generated Sheets</div></div>', unsafe_allow_html=True)

    st.markdown("---")
    st.subheader("How to Use")
    st.markdown("""
    1. **New Screening** — Enter SME loan application details and answer 7 eligibility questions
    2. **Automatic Calculation** — The tool counts 'No' answers and determines if the SME is satisfactory
    3. **Admin Dashboard** — View pipeline summary and all entries
    4. **Download Data** — Export all entries as an Excel file with formulas and dashboard
    """)

    st.markdown("---")
    st.subheader("Decision Rule")
    st.info(f"Currently configured: An SME must have **≤ {st.session_state.threshold} 'No' answer(s)** to be rated 'Satisfactory' and recommended for Stage 2.")

# New Screening Page
elif page == "📝 New Screening":
    st.markdown('<div class="main-header">New Screening Entry</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-header">Fill in the details below. Yellow fields are required inputs.</div>', unsafe_allow_html=True)

    with st.form("screening_form"):
        st.subheader("📋 Basic Information")
        col1, col2 = st.columns(2)
        with col1:
            screening_date = st.date_input("Screening Date", datetime.now())
            sme_name = st.text_input("SME Name *", placeholder="e.g., Himal Herbs Pvt. Ltd.")
            district = st.text_input("District / Location *", placeholder="e.g., Bhojpur")
        with col2:
            sector = st.selectbox("Sector *", SECTORS)
            loan_amount = st.number_input("Loan Amount Applied (NPR) *", min_value=0, step=10000, value=0)
            bank_name = st.selectbox("BDSP / Partner Bank Name *", BANKS)

        screener_name = st.text_input("Screener Name *", placeholder="e.g., R. Sharma")

        st.markdown("---")
        st.subheader("✅ Eligibility Questions (Answer Yes / No)")

        answers = {}
        cols = st.columns(2)
        for i, (key, question) in enumerate(ELIGIBILITY_QUESTIONS.items()):
            with cols[i % 2]:
                answers[key] = st.selectbox(question, ["Yes", "No"], key=f"elig_{key}")

        st.markdown("---")
        st.subheader("💡 Innovation Details")
        innovation_type = st.selectbox("Likely Innovation Type (initial guess)", INNOVATION_TYPES)
        remarks = st.text_area("Remarks", placeholder="Describe the innovation briefly...")

        submitted = st.form_submit_button("💾 Save Screening Entry")

        if submitted:
            if not sme_name or not district or loan_amount <= 0 or not screener_name:
                st.error("Please fill in all required fields (marked with *)")
            else:
                result, action, no_count = calculate_result(answers, st.session_state.threshold)

                entry = {
                    'screening_date': screening_date.strftime('%Y-%m-%d'),
                    'sme_name': sme_name,
                    'district': district,
                    'sector': sector,
                    'loan_amount': loan_amount,
                    'bank_name': bank_name,
                    'screener_name': screener_name,
                    **answers,
                    'innovation_type': innovation_type,
                    'remarks': remarks,
                    'result': result,
                    'action': action,
                    'no_count': no_count
                }

                st.session_state.entries.append(entry)

                st.success(f"✅ Entry saved! **{sme_name}** is rated: **{result}** — {action}")
                if result == "Satisfactory":
                    st.balloons()

# Admin Dashboard Page
elif page == "📊 Admin Dashboard":
    st.markdown('<div class="main-header">Admin Dashboard</div>', unsafe_allow_html=True)

    # Admin login
    if not st.session_state.admin_logged_in:
        with st.form("admin_login"):
            st.warning("🔒 Admin access required")
            password = st.text_input("Enter admin password", type="password")
            if st.form_submit_button("Login"):
                if password == "admin123":  # Simple password for demo
                    st.session_state.admin_logged_in = True
                    st.rerun()
                else:
                    st.error("Incorrect password")
    else:
        st.success("✅ Logged in as Administrator")
        if st.button("Logout"):
            st.session_state.admin_logged_in = False
            st.rerun()

        # Settings
        with st.expander("⚙️ Configure Decision Rule"):
            new_threshold = st.number_input(
                "Maximum number of 'No' answers still considered Satisfactory:",
                min_value=0, max_value=7, value=st.session_state.threshold
            )
            if st.button("Update Threshold"):
                st.session_state.threshold = new_threshold
                st.success(f"Threshold updated to {new_threshold}")

        st.markdown("---")

        # Stats
        total, satisfactory, non_satisfactory, pct, innovation_counts = get_dashboard_stats(
            st.session_state.entries, st.session_state.threshold
        )

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.markdown(f'<div class="metric-card"><div class="metric-value">{total}</div><div class="metric-label">Total Screened</div></div>', unsafe_allow_html=True)
        with col2:
            st.markdown(f'<div class="metric-card"><div class="metric-value" style="color:#28a745;">{satisfactory}</div><div class="metric-label">Satisfactory</div></div>', unsafe_allow_html=True)
        with col3:
            st.markdown(f'<div class="metric-card"><div class="metric-value" style="color:#dc3545;">{non_satisfactory}</div><div class="metric-label">Non-Satisfactory</div></div>', unsafe_allow_html=True)
        with col4:
            st.markdown(f'<div class="metric-card"><div class="metric-value">{pct:.1f}%</div><div class="metric-label">Recommended</div></div>', unsafe_allow_html=True)

        st.markdown("---")

        # Innovation type breakdown
        st.subheader("By Likely Innovation Type")
        innov_df = pd.DataFrame([
            {"Innovation Type": k, "Count": v} 
            for k, v in innovation_counts.items() if v > 0 or True
        ])
        st.bar_chart(innov_df.set_index("Innovation Type"))

        # All entries table
        st.markdown("---")
        st.subheader("All Screening Entries")

        if st.session_state.entries:
            df = pd.DataFrame(st.session_state.entries)
            df['SN'] = range(1, len(df) + 1)
            display_cols = ['SN', 'screening_date', 'sme_name', 'district', 'sector', 
                           'loan_amount', 'bank_name', 'screener_name', 'innovation_type', 
                           'result', 'action', 'no_count']

            # Style the dataframe
            def highlight_result(row):
                if row['result'] == 'Satisfactory':
                    return ['background-color: #d4edda'] * len(row)
                else:
                    return ['background-color: #f8d7da'] * len(row)

            styled_df = df[display_cols].style.apply(highlight_result, axis=1)
            st.dataframe(styled_df, use_container_width=True)
        else:
            st.info("No entries yet. Go to 'New Screening' to add entries.")

# Download Page
elif page == "📥 Download Data":
    st.markdown('<div class="main-header">Download Screening Data</div>', unsafe_allow_html=True)

    if not st.session_state.entries:
        st.warning("No entries to download. Please add screening entries first.")
    else:
        st.success(f"Ready to export **{len(st.session_state.entries)}** screening entries")

        # Preview
        with st.expander("👁️ Preview Data"):
            df = pd.DataFrame(st.session_state.entries)
            st.dataframe(df, use_container_width=True)

        # Generate Excel
        excel_file = create_excel_with_formulas(st.session_state.entries, st.session_state.threshold)

        st.download_button(
            label="📥 Download Excel File (with formulas & dashboard)",
            data=excel_file,
            file_name=f"Tool1_Screening_Register_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.info("The downloaded Excel file includes:

"
                "• **Instructions** sheet with configurable threshold
"
                "• **Screening Register** with all entries and automatic formulas
"
                "• **Dashboard** with pipeline summary that auto-updates")

        # Also offer CSV
        csv_df = pd.DataFrame(st.session_state.entries)
        csv = csv_df.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="📄 Download CSV (raw data only)",
            data=csv,
            file_name=f"Tool1_Screening_Data_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv"
        )
